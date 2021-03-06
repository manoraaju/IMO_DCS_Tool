﻿## Import Libraries

import tkinter as tk
import threading, numbers, copy
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter.font import Font
import pandas as pd
from datetime import datetime, timedelta
import os, openpyxl, sys, math
import numpy as np
import openpyxl.styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils.dataframe import dataframe_to_rows
import joblib
from tensorflow.keras.models import load_model
from collections import OrderedDict
from PyPDF2 import PdfFileWriter, PdfFileReader
from PyPDF2.generic import BooleanObject, NameObject, IndirectObject
from pprint import pprint
## Global Variables
global country_codes ## ISO3166, Alpha 3
country_codes = {'Afghanistan':'AFG',
'Albania':'ALB',
'Algeria':'DZA',
'American Samoa':'ASM',
'Andorra':'AND',
'Angola':'AGO',
'Anguilla':'AIA',
'Antarctica':'ATA',
'Antigua and Barbuda':'ATG',
'Argentina':'ARG',
'Armenia':'ARM',
'Aruba':'ABW',
'Australia':'AUS',
'Austria':'AUT',
'Azerbaijan':'AZE',
'Bahamas (the)':'BHS',
'Bahrain':'BHR',
'Bangladesh':'BGD',
'Barbados':'BRB',
'Belarus':'BLR',
'Belgium':'BEL',
'Belize':'BLZ',
'Benin':'BEN',
'Bermuda':'BMU',
'Bhutan':'BTN',
'Bolivia (Plurinational State of)':'BOL',
'Bonaire, Sint Eustatius and Saba':'BES',
'Bosnia and Herzegovina':'BIH',
'Botswana':'BWA',
'Bouvet Island':'BVT',
'Brazil':'BRA',
'British Indian Ocean Territory (the)':'IOT',
'Brunei Darussalam':'BRN',
'Bulgaria':'BGR',
'Burkina Faso':'BFA',
'Burundi':'BDI',
'Cabo Verde':'CPV',
'Cambodia':'KHM',
'Cameroon':'CMR',
'Canada':'CAN',
'Cayman Islands (the)':'CYM',
'Central African Republic (the)':'CAF',
'Chad':'TCD',
'Chile':'CHL',
'China':'CHN',
'Christmas Island':'CXR',
'Cocos (Keeling) Islands (the)':'CCK',
'Colombia':'COL',
'Comoros (the)':'COM',
'Congo (the Democratic Republic of the)':'COD',
'Congo (the)':'COG',
'Cook Islands (the)':'COK',
'Costa Rica':'CRI',
'Croatia':'HRV',
'Cuba':'CUB',
'Curaçao':'CUW',
'Cyprus':'CYP',
'Czechia':'CZE',
"Côte d'Ivoire":'CIV',
'Denmark':'DNK',
'Djibouti':'DJI',
'Dominica':'DMA',
'Dominican Republic (the)':'DOM',
'Ecuador':'ECU',
'Egypt':'EGY',
'El Salvador':'SLV',
'Equatorial Guinea':'GNQ',
'Eritrea':'ERI',
'Estonia':'EST',
'Eswatini':'SWZ',
'Ethiopia':'ETH',
'Falkland Islands (the) [Malvinas]':'FLK',
'Faroe Islands (the)':'FRO',
'Fiji':'FJI',
'Finland':'FIN',
'France':'FRA',
'French Guiana':'GUF',
'French Polynesia':'PYF',
'French Southern Territories (the)':'ATF',
'Gabon':'GAB',
'Gambia (the)':'GMB',
'Georgia':'GEO',
'Germany':'DEU',
'Ghana':'GHA',
'Gibraltar':'GIB',
'Greece':'GRC',
'Greenland':'GRL',
'Grenada':'GRD',
'Guadeloupe':'GLP',
'Guam':'GUM',
'Guatemala':'GTM',
'Guernsey':'GGY',
'Guinea':'GIN',
'Guinea-Bissau':'GNB',
'Guyana':'GUY',
'Haiti':'HTI',
'Heard Island and McDonald Islands':'HMD',
'Holy See (the)':'VAT',
'Honduras':'HND',
'Hong Kong':'HKG',
'Hungary':'HUN',
'Iceland':'ISL',
'India':'IND',
'Indonesia':'IDN',
'Iran (Islamic Republic of)':'IRN',
'Iraq':'IRQ',
'Ireland':'IRL',
'Isle of Man':'IMN',
'Israel':'ISR',
'Italy':'ITA',
'Jamaica':'JAM',
'Japan':'JPN',
'Jersey':'JEY',
'Jordan':'JOR',
'Kazakhstan':'KAZ',
'Kenya':'KEN',
'Kiribati':'KIR',
"Korea (the Democratic People's Republic of)":'PRK',
'Korea (the Republic of)':'KOR',
'Kuwait':'KWT',
'Kyrgyzstan':'KGZ',
"Lao People's Democratic Republic (the)":'LAO',
'Latvia':'LVA',
'Lebanon':'LBN',
'Lesotho':'LSO',
'Liberia':'LBR',
'Libya':'LBY',
'Liechtenstein':'LIE',
'Lithuania':'LTU',
'Luxembourg':'LUX',
'Macao':'MAC',
'Madagascar':'MDG',
'Malawi':'MWI',
'Malaysia':'MYS',
'Maldives':'MDV',
'Mali':'MLI',
'Malta':'MLT',
'Marshall Islands (the)':'MHL',
'Martinique':'MTQ',
'Mauritania':'MRT',
'Mauritius':'MUS',
'Mayotte':'MYT',
'Mexico':'MEX',
'Micronesia (Federated States of)':'FSM',
'Moldova (the Republic of)':'MDA',
'Monaco':'MCO',
'Mongolia':'MNG',
'Montenegro':'MNE',
'Montserrat':'MSR',
'Morocco':'MAR',
'Mozambique':'MOZ',
'Myanmar':'MMR',
'Namibia':'NAM',
'Nauru':'NRU',
'Nepal':'NPL',
'Netherlands (the)':'NLD',
'New Caledonia':'NCL',
'New Zealand':'NZL',
'Nicaragua':'NIC',
'Niger (the)':'NER',
'Nigeria':'NGA',
'Niue':'NIU',
'Norfolk Island':'NFK',
'Northern Mariana Islands (the)':'MNP',
'Norway':'NOR',
'Oman':'OMN',
'Pakistan':'PAK',
'Palau':'PLW',
'Palestine, State of':'PSE',
'Panama':'PAN',
'Papua New Guinea':'PNG',
'Paraguay':'PRY',
'Peru':'PER',
'Philippines (the)':'PHL',
'Pitcairn':'PCN',
'Poland':'POL',
'Portugal':'PRT',
'Puerto Rico':'PRI',
'Qatar':'QAT',
'Republic of North Macedonia':'MKD',
'Romania':'ROU',
'Russian Federation (the)':'RUS',
'Rwanda':'RWA',
'Réunion':'REU',
'Saint Barthélemy':'BLM',
'Saint Helena, Ascension and Tristan da Cunha':'SHN',
'Saint Kitts and Nevis':'KNA',
'Saint Lucia':'LCA',
'Saint Martin (French part)':'MAF',
'Saint Pierre and Miquelon':'SPM',
'Saint Vincent and the Grenadines':'VCT',
'Samoa':'WSM',
'San Marino':'SMR',
'Sao Tome and Principe':'STP',
'Saudi Arabia':'SAU',
'Senegal':'SEN',
'Serbia':'SRB',
'Seychelles':'SYC',
'Sierra Leone':'SLE',
'Singapore':'SGP',
'Sint Maarten (Dutch part)':'SXM',
'Slovakia':'SVK',
'Slovenia':'SVN',
'Solomon Islands':'SLB',
'Somalia':'SOM',
'South Africa':'ZAF',
'South Georgia and the South Sandwich Islands':'SGS',
'South Sudan':'SSD',
'Spain':'ESP',
'Sri Lanka':'LKA',
'Sudan (the)':'SDN',
'Suriname':'SUR',
'Svalbard and Jan Mayen':'SJM',
'Sweden':'SWE',
'Switzerland':'CHE',
'Syrian Arab Republic':'SYR',
'Taiwan (Province of China)':'TWN',
'Tajikistan':'TJK',
'Tanzania, United Republic of':'TZA',
'Thailand':'THA',
'Timor-Leste':'TLS',
'Togo':'TGO',
'Tokelau':'TKL',
'Tonga':'TON',
'Trinidad and Tobago':'TTO',
'Tunisia':'TUN',
'Turkey':'TUR',
'Turkmenistan':'TKM',
'Turks and Caicos Islands (the)':'TCA',
'Tuvalu':'TUV',
'Uganda':'UGA',
'Ukraine':'UKR',
'United Arab Emirates (the)':'ARE',
'United Kingdom of Great Britain and Northern Ireland (the)':'GBR',
'United States Minor Outlying Islands (the)':'UMI',
'United States of America (the)':'USA',
'Uruguay':'URY',
'Uzbekistan':'UZB',
'Vanuatu':'VUT',
'Venezuela (Bolivarian Republic of)':'VEN',
'Viet Nam':'VNM',
'Virgin Islands (British)':'VGB',
'Virgin Islands (U.S.)':'VIR',
'Wallis and Futuna':'WLF',
'Western Sahara':'ESH',
'Yemen':'YEM',
'Zambia':'ZMB',
'Zimbabwe':'ZWE',
'Åland Islands':'ALA'}
global dcs_template_columns
dcs_template_columns = ['Start Date and Time (DD-Mmm-YYYY HH:MM) (UTC)*',
                        'End Date and Time (DD-Mmm-YYYY HH:MM) (UTC)*', 'Hours Underway (h)*',
                        'Distance Traveled (nm)*', 'HFO Consumption (MT)',
                        'LFO Consumption (MT)', 'Diesel/Gas Oil Consumption (MT)',
                        'LPG (Propane) Consumption (MT)', 'LPG (Butane) Consumption (MT)',
                        'LNG Consumption (MT)', 'Methanol Consumption (MT)',
                        'Ethanol Consumption (MT)', 'Other Consumption (MT)',
                        'Other Emission Factor (MT CO2/MT fuel)',
                        'Direct CO2 emissions measurement (MT), if applicable']

global dcs_numeric_columns
dcs_numeric_columns = dcs_template_columns[2:]

global dcs_fuel_columns
dcs_fuel_columns = dcs_template_columns[4:13]

global bdn_template_columns
bdn_template_columns = ["Entry Type*", "Date (DD-Mmm-YYYY)*", "HFO (MT)", "LFO (MT)", "Diesel/Gas Oil (MT)",
                        "LPG (Propane) (MT)", "LPG (Butane) (MT)", "LNG (MT)", "Methanol (MT)", "Ethanol (MT)",
                        "Other (MT)", "Other Emission Factor (MT CO2/MT fuel)"]

global bdn_entry_type
bdn_entry_type = ["BDN", "ROB (Start)", "ROB (End)",
                  "Other Corrections"]

global fuel_mapping
fuel_mapping = {"HFO (MT)": "HFO Consumption (MT)",
                "LFO (MT)": "LFO Consumption (MT)",
                "Diesel/Gas Oil (MT)": "Diesel/Gas Oil Consumption (MT)",
                "LPG (Propane) (MT)": "LPG (Propane) Consumption (MT)",
                "LPG (Butane) (MT)": "LPG (Butane) Consumption (MT)",
                "LNG (MT)": "LNG Consumption (MT)",
                "Methanol (MT)": "Methanol Consumption (MT)",
                "Ethanol (MT)": "Ethanol Consumption (MT)",
                "Other (MT)": "Other Consumption (MT)"}

global bdn_numeric_columns
bdn_numeric_columns = bdn_template_columns[2:]

global bdn_fuel_columns
bdn_fuel_columns = bdn_template_columns[2:11]

global start_year_date
start_year_date = datetime.strptime("1-Jan-2019", "%d-%b-%Y")

global end_year_date
if start_year_date.year % 4 == 0:
    no_of_days = 366
else:
    no_of_days = 365
end_year_date = start_year_date + timedelta(days=no_of_days) - timedelta(minutes=1)

global dcs_data_skip_row
dcs_data_skip_row = 6

global dcs_data_end_col
dcs_data_end_col = 16

global bdn_data_skip_row
bdn_data_skip_row = 6

global bdn_data_end_col
bdn_data_end_col = 13

global report_setup_columns
report_setup_columns = ["Field Description", "Client input", "Instructions"]

global rsetup_skip_row
rsetup_skip_row = 2

global rsetup_end_col
rsetup_end_col = 4

global rsetup_client_input
rsetup_client_input = ["End of Year", "Change of Flag", "Change of Company", "Change of Flag and Company",
                       "Decommission"]

global vessel_details_columns
vessel_details_columns = ["Field Description", "Client input", "Instructions"]

global vdetails_skip_row
vdetails_skip_row = 3

global vdetails_end_col
vdetails_end_col = 4

## Fill patterns declaration
global error_yellow_fill
error_yellow_fill = PatternFill("solid", fgColor="00DDDD00")

class TextRedirector(object):
    def __init__(self, widget, tag="stdout"):
        self.widget = widget
        self.tag = tag

    def write(self, str):
        self.widget.configure(state="normal")
        self.widget.insert("end", str, (self.tag,))
        self.widget.configure(state="disabled")


class IMO_DCS_App(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.title("IMO DCS Data Validator")
        self.maxsize(700, 600)  # width x height
        self.config(bg="skyblue")
        Label_Font = Font(family="Arial", size=10)
        top_frame = tk.Frame(self, width=200, height=200, bg='grey')
        top_frame.pack(side=TOP)

        self.load_folderPath = StringVar()
        self.output_folderPath = StringVar()
        ## Loading Excel Folder
        f_path = Label(top_frame, text="Enter Folder Path", font=Label_Font, bg='grey')
        f_path.grid(row=0, column=0, padx=5, pady=5, sticky='w')
        load_path_entry = Entry(top_frame, textvariable=self.load_folderPath, width=30)
        load_path_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w' + 'e' + 'n' + 's')
        btn_Find = ttk.Button(top_frame, text="Browse Folder", command=self.getFolderPath)
        btn_Find.grid(row=0, column=2, padx=5, pady=5, sticky='w' + 'e' + 'n' + 's')

        ## Selecting Output Folder
        o_path = Label(top_frame, text="Enter Output Path", font=Label_Font, bg='grey')
        o_path.grid(row=1, column=0, padx=5, pady=5, sticky='w')
        load_path_entry = Entry(top_frame, textvariable=self.output_folderPath, width=30)
        load_path_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w' + 'e' + 'n' + 's')
        btn_Find = ttk.Button(top_frame, text="Browse Folder", command=self.setFolderPath)
        btn_Find.grid(row=1, column=2, padx=5, pady=5, sticky='w' + 'e' + 'n' + 's')
        self.mistatement = tk.IntVar(self)
        self.cbox1 = tk.Checkbutton(top_frame, text="Misstatement",
                                    variable=self.mistatement, bg='grey', command=self.disable_GISIS)
        self.cbox1.grid(row=2, column=0, padx=5, pady=5, sticky='w')

        self.sampling = tk.IntVar(self)
        self.cbox2 = tk.Checkbutton(top_frame, text="Sampling",
                                    variable=self.sampling, bg='grey', command=self.disable_GISIS)
        self.cbox2.grid(row=2, column=1, padx=5, pady=5, sticky='w' + 'e' + 'n' + 's')

        self.GISIS = tk.IntVar(self)
        self.cbox3 = tk.Checkbutton(top_frame, text="GISIS Data-Gen",
                                    variable=self.GISIS, bg='grey', state=NORMAL, command=self.disable_mistatement_sampling)
        self.cbox3.grid(row=2, column=2, padx=5, pady=5, sticky='w')

        self.SOC = tk.IntVar(self)
        self.cbox4 = tk.Checkbutton(top_frame, text="SoC Gen",
                                    variable=self.SOC, bg='grey', state=NORMAL, command=self.disable_mis_GISIS)
        self.cbox4.grid(row=3, column=0, padx=5, pady=5, sticky='w')

        self.RR = tk.IntVar(self)
        self.cbox5 = tk.Checkbutton(top_frame, text="RR Gen",
                                    variable=self.RR, bg='grey', state=NORMAL, command=self.disable_mis_GISIS)
        self.cbox5.grid(row=3, column=1, padx=5, pady=5, sticky='w' + 'e' + 'n' + 's')

        self.btn_Start = ttk.Button(top_frame, text="Start", command=self.begin)
        self.btn_Start.grid(row=3, column=2, padx=5, pady=5, sticky=  'n' + 's')

        bottom_frame = tk.Frame(self, bg='grey')
        bottom_frame.pack(side=BOTTOM)

        self.text = tk.Text(bottom_frame, wrap="word")
        self.text.grid(row=0, column=0, padx=10, pady=5)
        self.text.config(state=DISABLED)
        # create a Scrollbar and associate it with txt
        scrollb = ttk.Scrollbar(bottom_frame, command=self.text.yview)
        scrollb.grid(row=0, column=1, sticky='nsew')
        self.text['yscrollcommand'] = scrollb.set

    def disable_mistatement_sampling(self):
        if self.GISIS.get():
            self.cbox1.config(state=DISABLED)
            self.cbox2.config(state=DISABLED)
            self.cbox4.config(state=DISABLED)
            self.cbox5.config(state=DISABLED)
        else:
            self.cbox1.config(state=NORMAL)
            self.cbox2.config(state=NORMAL)
            self.cbox4.config(state=NORMAL)
            self.cbox5.config(state=NORMAL)

    def disable_GISIS(self):
        if self.mistatement.get() or self.sampling.get():
            self.cbox3.config(state=DISABLED)
            self.cbox4.config(state=DISABLED)
            self.cbox5.config(state=DISABLED)
        else:
            self.cbox3.config(state=NORMAL)
            self.cbox4.config(state=NORMAL)
            self.cbox5.config(state=NORMAL)

    def disable_mis_GISIS(self):
        if self.SOC.get() or self.RR.get():
            self.cbox1.config(state=DISABLED)
            self.cbox2.config(state=DISABLED)
            self.cbox3.config(state=DISABLED)
        else:
            self.cbox1.config(state=NORMAL)
            self.cbox2.config(state=NORMAL)
            self.cbox3.config(state=NORMAL)

    def begin(self):
        '''start a thread and connect it to func'''
        threading.Thread(target=self.folderSelection, daemon=True).start()

    def message_box(self, message):
        #print(message)
        message = message + "\n"
        self.text.tag_config('warning:', background="light yellow", foreground="blue2")
        self.text.tag_config('error:', background="yellow", foreground="red3")
        self.text.tag_config('ok:', background="pale green", foreground="blue2")
        self.text.tag_config('normal', foreground="gray15")
        if "warning:" in message.lower():
            self.text.config(state=NORMAL)
            self.text.insert("end", message, "warning:")
            self.text.config(state=DISABLED)
        elif "error:" in message.lower():
            self.text.config(state=NORMAL)
            self.text.insert("end", message, "error:")
            self.text.config(state=DISABLED)
        elif "ok:" in message.lower():
            self.text.config(state=NORMAL)
            self.text.insert("end", message, "ok:")
            self.text.config(state=DISABLED)
        else:
            self.text.config(state=NORMAL)
            self.text.insert("end", message, "normal")
            self.text.config(state=DISABLED)
        self.text.see("end")

    def check_date_range(self, date):
        if date >= start_year_date:
            if date <= end_year_date:
                return True
        return False

    def date_format(self, date):
        return datetime.strftime(date, "%d-%b-%Y")

    def is_NaT_NaN(self, n):
        if pd.isnull(n):
            return True
        if isinstance(n, float) and math.isnan(n):
            return True
        return False

    def is_numeric(self, n):
        try:
            if isinstance(n, int):
                return True
            if isinstance(n, float):
                return True
            return False
        except:
            return False

    def is_nonnegative(self, n):
        try:
            if isinstance(n, int):
                if int(n) >= 0:
                    return True
            if isinstance(n, float):
                if float(n) >= 0:
                    return True
            return False
        except:
            return False

    def as_text(self, value):
        if value is None:
            return ""
        return str(value)

    def fill_cell(self, df, index, col_name, pattern, sheet_name):
        if sheet_name == "IMO DCS Data":
            skip_rows = dcs_data_skip_row
            col_num = 1
        elif sheet_name == "IMO DCS BDN Summary":
            skip_rows = bdn_data_skip_row
            col_num = 1
        elif sheet_name == "Report Setup":
            skip_rows = rsetup_skip_row
            col_num = 1
        elif sheet_name == "Vessel Details":
            skip_rows = vdetails_skip_row
            col_num = 1
        elif sheet_name == "Program Checklist":
            skip_rows = 0
            col_num = 2
        ws = wb[sheet_name]
        ws.cell(row=(index + skip_rows + 2), column=df.columns.get_loc(col_name) + col_num).fill = error_yellow_fill

    def add_column_openpyxl(self, column, sheet_name):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ws = wb[sheet_name]
        new_column = ws.max_column + 1
        if sheet_name == "IMO DCS Data":
            skip_rows = dcs_data_skip_row
            new_column = 17
        elif sheet_name == "IMO DCS BDN Summary":
            skip_rows = bdn_data_skip_row
            new_column = 14
        elif sheet_name == "Report Setup":
            skip_rows = rsetup_skip_row
            new_column = 21
        elif sheet_name == "Vessel Details":
            skip_rows = vdetails_skip_row
            new_column = 21
        elif sheet_name == "Program Checklist":
            skip_rows = 0
        for rowy, value in enumerate(column, start=1):
            ws.cell(row=rowy + skip_rows, column=new_column, value=value)
            ws.cell(row=rowy + skip_rows, column=new_column, ).border = thin_border

    def getFolderPath(self):
        folder_selected = filedialog.askdirectory()
        self.load_folderPath.set(folder_selected)

    def setFolderPath(self):
        folder_selected = filedialog.askdirectory()
        self.output_folderPath.set(folder_selected)

    def folderSelection(self):
        input_folder = self.load_folderPath.get()
        output_folder = self.output_folderPath.get()
        try:
            if len(input_folder) == 0 or len(output_folder) == 0:
                raise
            else:
                msg = "Message: The selected input folder is {}".format(input_folder)
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
                msg = ("Message: The selected output folder is {}".format(output_folder))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            if len(input_folder) == 0 or len(output_folder) == 0:
                msg = ("Error: One of the folders is not selected! Try again!")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
            else:
                msg = ("Message: There is no folder selected! Please select the input and output folder")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        else:
            self.check_files(input_folder, output_folder)

    def check_files(self, s_dir, d_dir):
        ## Main start
        source_dir = s_dir
        destination_dir = d_dir

        ## Check for all the xlsx files in the source directory
        files = [filename for filename in os.listdir(source_dir)]
        files = [file for file in files if ((file.lower()).endswith(".xlsx") or (file.lower()).endswith(".xls"))]
        files = [file for file in files if file[0:2] != "~$"]
        files = [file for file in files if file[0:7] != "Output_"]

        if len(files) == 0:
            msg = ("Warning: The selected folder do not contain spreadsheet.")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
        else:
            msg = ("OK: The selected folder contains files: \n{}".format(";\n".join(files)))
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            if (not self.sampling.get()) and (not self.mistatement.get()) and (not self.GISIS.get())\
                    and (not self.SOC.get()) and (not self.RR.get()):
                msg = ("Error: Please select one of the check box")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
            else:
                self.main_program(source_dir, destination_dir, files)


    def misstatement_prediction(self, source_dir, f_name):
        df_analysis = pd.DataFrame(columns=bdn_fuel_columns)
        df_analysis_errors = []
        df_errors = {"Vessel Details": [], "Report Setup": [], "IMO DCS BDN Summary": [], "IMO DCS Data": []}
        ## ===============================================================================================================
        ## For analysis of Vessel Details
        msg = ("Analysis started for file: {} and sheet name {}".format(f_name, "Vessel Details"))
        msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
        self.message_box(msg)
        vessel_details_flag = False
        vessel_details_errors = []
        filename = source_dir + "\\" + f_name
        ## Import file
        df = pd.DataFrame()
        vessel_details_flag = True
        vessel_types = ["Bulk Carrier", "Gas Carrier", "Tanker", "Containerships", "General Cargo Ship",
                        "Refrigerated Cargo Carrier", "Combination Carrier", "Passenger Ship",
                        "Ro-ro Cargo Ship (Vehicle Carrier)", "Ro-ro Cargo Ship", "Ro-ro Passenger Ship",
                        "LNG Carrier", "Cruise Passenger Ship", "Others"]
        try:
            df = pd.read_excel(filename, sheet_name='Vessel Details', skiprows=vdetails_skip_row,
                               usecols=range(vdetails_end_col), keep_default_na=False, na_values=[''])
            msg = ("OK: Importing sheet of Vessel Details")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            vessel_details_flag = True
        except:
            msg = ("Critical Error: Importing sheet of Vessel Details")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            vessel_details_flag = False
            vessel_details_errors.append(msg)

        df["Errors"] = ""
        ## Check the imported file or template
        vdetails_missing_columns = []
        for col in report_setup_columns:
            if col not in df.columns:
                vdetails_missing_columns.append(col)

        if len(vdetails_missing_columns) != 0:
            msg = ("Critical Error: Template Missing columns: {}".format(",".join(vdetails_missing_columns)))
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            vessel_details_flag = False
            vessel_details_errors.append(msg)
        else:
            msg = ("OK: Template")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            vessel_details_flag = True

        try:
            if vessel_details_flag:
                for index, row in df.iterrows():
                    temp_error_msgs = []
                    if self.is_NaT_NaN(row["Client input"]):
                        ship_type = wb['Vessel Details']["B7"].value
                        print(index)
                        if index == 3 and ship_type != "Others":
                            continue
                        if index == 3 and ship_type == "Others":
                            temp_error_msgs.append("'Other' vessel type selected. Please enter description;")
                            self.fill_cell(df, index, "Client input", error_yellow_fill, "Vessel Details")
                        if index == 2 and ship_type not in vessel_types:
                            temp_error_msgs.append("Please select vessel type from dropdown list;")
                            self.fill_cell(df, index, "Client input", error_yellow_fill, "Vessel Details")
                        else:
                            temp_error_msgs.append("Empty cell. Please enter data;")
                            self.fill_cell(df, index, "Client input", error_yellow_fill, "Vessel Details")
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check Client Input type")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check Input type")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            vessel_details_errors.append(msg)

        #print(df.iloc[2]["Client input"])
        '''
        try:
            ## Check input type
            if vessel_details_flag:
                temp_error_msgs = []
                if df.iloc[2]["Client input"] not in vessel_types:
                    temp_error_msgs.append("Please select vessel type from dropdown list;")
                    self.fill_cell(df, 2, "Client input", error_yellow_fill, "Vessel Details")
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[2, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check Vessel Input type")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check Vessel Input type")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            vessel_details_errors.append(msg)
        '''

        try:
            if vessel_details_flag:
                df_errors["Vessel Details"] = df["Errors"]
                error_column = ["Errors"]
                error_column.extend(df["Errors"])
                df_errors["Vessel Details"] = df["Errors"]
                self.add_column_openpyxl(error_column, "Vessel Details")
                msg = ("OK: Appending 'Error' column into Vessel Details sheet for file {}".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
                msg = ("OK: Analysis of Vessel Details Successful for file {}".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Appending 'Error' column into Vessel Details sheet for file {}".format(f_name))
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)

        ## ===============================================================================================================
        ## For analysis of Report Setup
        msg = ("Analysis started for file: {} and sheet name {}".format(f_name, "Report Setup"))
        msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
        self.message_box(msg)
        report_setup_flag = False
        report_setup_errors = []
        filename = source_dir + "\\" + f_name
        ## Import file
        df = pd.DataFrame()
        try:
            df = pd.read_excel(filename, sheet_name='Report Setup', skiprows=rsetup_skip_row,
                               usecols=range(rsetup_end_col), keep_default_na=False,na_values=[''])
            msg = ("OK: Importing sheet of Report Setup")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            report_setup_flag = True
        except:
            msg = ("Critical Error: Importing sheet of Report Setup")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            report_setup_flag = False
            report_setup_errors.append(msg)

        df["Errors"] = ""
        ## Check the imported file or template
        rsetup_missing_columns = []
        for col in report_setup_columns:
            if col not in df.columns:
                rsetup_missing_columns.append(col)

        if len(rsetup_missing_columns) != 0:
            msg = ("Critical Error: Template Missing columns: {}".format(",".join(rsetup_missing_columns)))
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            report_setup_flag = False
            report_setup_errors.append(msg)
        else:
            msg = ("OK: Template")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            report_setup_flag = True

        try:
            ## Check input type
            if report_setup_flag:
                EoY_flag = False
                CoM_flag = False
                DeCom_flag = False
                for index, row in df.iterrows():
                    temp_error_msgs = []
                    col_name = "Client input"
                    if not (row.isnull().all()):

                        if self.is_NaT_NaN(row[col_name]):
                            temp_error_msgs.append("Please select a value from the drop-down list under "
                                                   "'Entry Type' column in the template.")
                            self.fill_cell(df, index, col_name, error_yellow_fill, "Report Setup")
                        elif row[col_name] not in rsetup_client_input:
                            temp_error_msgs.append("Please select a value from the drop-down list under "
                                                   "'Entry Type' column in the template.")
                            self.fill_cell(df, index, col_name, error_yellow_fill, "Report Setup")
                        if row[col_name] == "End of Year":
                            EoY_flag = True
                        elif row[col_name] in ["Change of Flag and Company", "Change of Flag", "Change of Company"]:
                            CoM_flag = True
                        elif row[col_name] == "Decommission":
                            DeCom_flag = True
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check Client Input type")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check Input type")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            report_setup_errors.append(msg)

        try:
            if report_setup_flag:
                error_column = ["Errors"]
                error_column.extend(df["Errors"])
                df_errors["Report Setup"] = df["Errors"]
                self.add_column_openpyxl(error_column, "Report Setup")
                msg = ("OK: Appending 'Error' column into Report Setup sheet for file {}".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
                msg = ("OK: Analysis of Report Setup Successful for file {}".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Appending 'Error' column into Report Setup sheet for file {}".format(f_name))
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            report_setup_errors.append(msg)

        ## =================================================================================================================
        ## For analysing IMO DCS BDN Summary
        msg = ("Analysis started for file: {} and sheet name {}".format(f_name, "IMO DCS BDN Summary"))
        msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
        self.message_box(msg)
        dcs_bdn_flag = False
        dcs_bdn_errors = []
        filename = source_dir + "\\" + f_name
        ## Import file
        df = pd.DataFrame()
        try:
            df = pd.read_excel(filename, sheet_name='IMO DCS BDN Summary', skiprows=bdn_data_skip_row,
                               usecols=range(bdn_data_end_col), keep_default_na=False,na_values=[''])
            msg = ("OK: Importing sheet of IMO DCS BDN Summary")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_bdn_flag = True
        except:
            msg = ("Critical Error: Importing sheet of IMO DCS BDN Summary")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_bdn_flag = False
            dcs_bdn_errors.append(msg)

        df["Errors"] = ""

        ## Check the imported file or template
        bdn_missing_columns = []
        for col in bdn_template_columns:
            if col not in df.columns:
                bdn_missing_columns.append(col)

        if len(bdn_missing_columns) != 0:
            msg = ("Critical Error: Template Missing columns: {}".format(",".join(bdn_missing_columns)))
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_bdn_flag = False
            dcs_bdn_errors.append(msg)
        else:
            msg = ("OK: Template")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_bdn_flag = True

        try:
            ## Check entry type
            if dcs_bdn_flag:
                for index, row in df.iterrows():
                    temp_error_msgs = []
                    if not (row.isnull().all()):
                        if self.is_NaT_NaN(row["Entry Type*"]):
                            temp_error_msgs.append("Please select a value from the drop-down list under 'Entry Type' "
                                                   "column in the template;")
                            self.fill_cell(df, index, "Entry Type*", error_yellow_fill, "IMO DCS BDN Summary")
                        elif (row["Entry Type*"]).strip() not in bdn_entry_type:
                            temp_error_msgs.append("Please select a value from the drop-down list under 'Entry Type' "
                                                   "column in the template;")
                            self.fill_cell(df, index, "Entry Type*", error_yellow_fill, "IMO DCS BDN Summary")
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check entry type")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check entry type")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_bdn_errors.append(msg)

        try:
            ## Check if numeric in columns
            if dcs_bdn_flag:
                for index, row in df.iterrows():
                    temp_error_msgs = []
                    non_numeric_colnames = []
                    if not (row.isnull().all()):
                        for col_name in bdn_numeric_columns:
                            if not self.is_NaT_NaN(row[col_name]):
                                if not self.is_numeric(row[col_name]):  ## Check if the data is numeric
                                    non_numeric_colnames.append(col_name)
                                    self.fill_cell(df, index, col_name, error_yellow_fill, "IMO DCS BDN Summary")
                    if len(non_numeric_colnames) > 0:
                        temp_error_msgs.append("Non-Numeric Value in Column(s): {};"
                                               .format(",".join(non_numeric_colnames)))
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check if numeric in columns")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check if numeric in columns")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_bdn_errors.append(msg)

        try:
            ## Check for non-zero in Other Consumption (MT) and Emission Factor
            if dcs_bdn_flag:
                for index, row in df.iterrows():
                    other_consump = row["Other (MT)"]
                    other_emission = row["Other Emission Factor (MT CO2/MT fuel)"]
                    temp_error_msgs = []
                    if not (row.isnull().all()):
                        if self.is_numeric(other_consump) and self.is_numeric(other_emission):
                            if abs(other_consump) > 0 and (other_emission == 0 or self.is_NaT_NaN(other_emission)):
                                temp_error_msgs.append(
                                    "When a value for “Other (MT)” is entered, "
                                    "a value for “Other Emission Factor” is to be entered."
                                    "Please check your data;")
                                self.fill_cell(df, index, "Other (MT)", error_yellow_fill, "IMO DCS BDN Summary")
                            elif (other_consump == 0 or self.is_NaT_NaN(other_consump)) and abs(other_emission) > 0:
                                temp_error_msgs.append(
                                    "When a value for “Other Emission Factor” is entered, "
                                    "a value for “Other (MT)” is to be entered."
                                    "Please check your data;")
                                self.fill_cell(df, index, "Other Emission Factor (MT CO2/MT fuel)", error_yellow_fill,
                                               "IMO DCS BDN Summary")
                            # if other_emission < 0:
                            #     temp_error_msgs.append(
                            #         "Negative value for “Other Emission Factor” is entered,"
                            #         "Please check your data;")
                            #     self.fill_cell(df, index, "Other Emission Factor (MT CO2/MT fuel)", error_yellow_fill,
                            #                    "IMO DCS BDN Summary")
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check for non-zero in Other Consumption (MT) and Emission Factor")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check for non-zero in Other Consumption (MT) and Emission Factor")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_bdn_errors.append(msg)


        try:
            ## Check if non-negative in columns
            if dcs_bdn_flag:
                for index, row in df.iterrows():
                    temp_error_msgs = []
                    negative_val_colnames = []
                    if not (row.isnull().all()):
                        for col_name in bdn_numeric_columns:
                            if not self.is_NaT_NaN(row[col_name]) and self.is_numeric(row[col_name]):
                                if not self.is_nonnegative(row[col_name]):  ## Check if the data is non-negative
                                    negative_val_colnames.append(col_name)
                                    self.fill_cell(df, index, col_name, error_yellow_fill, "IMO DCS BDN Summary")
                    if len(negative_val_colnames) > 0:
                        temp_error_msgs.append("Negative Value in Column(s): {};"
                                               .format(",".join(negative_val_colnames)))
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check if non-negative in columns")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check if non-negative in columns")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)

        try:
            ## Check for date format and empty date
            if dcs_bdn_flag:
                date_colname = "Date (DD-Mmm-YYYY)*"
                for index, row in df.iterrows():
                    temp_error_msgs = []
                    if not (row.isnull().all()):
                        if not isinstance(row[date_colname], datetime):
                            temp_error_msgs.append("Error in start date/time format (i.e. DD-Mmm-YYYY). "
                                                   "Please check data;")
                            self.fill_cell(df, index, date_colname, error_yellow_fill, "IMO DCS BDN Summary")
                        if isinstance(row[date_colname], datetime) and self.is_NaT_NaN(row[date_colname]):
                            temp_error_msgs.append("Empty cell in date/time. Please check data;")
                            self.fill_cell(df, index, date_colname, error_yellow_fill, "IMO DCS BDN Summary")
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check for date format")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check for date format")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_bdn_errors.append(msg)

        ## Check and get "ROB Correction (Start) and (End)" Testing phase...
        try:
            if dcs_bdn_flag:
                rob_start_cnt = 0
                rob_end_cnt = 0
                for index, row in df.iterrows():
                    temp_error_msgs = []
                    if row["Entry Type*"] == "ROB (Start)":
                        rob_start_cnt = rob_start_cnt + 1
                    if row["Entry Type*"] == "ROB (End)":
                        rob_end_cnt = rob_end_cnt + 1
                    if rob_start_cnt == 1 and (row["Entry Type*"] == "ROB (Start)"):
                        rob_start_date_val = row["Date (DD-Mmm-YYYY)*"]
                        if isinstance(rob_start_date_val, datetime) and (not self.is_NaT_NaN(rob_start_date_val)):
                            global start_year_date
                            start_year_date = rob_start_date_val
                        else:
                            temp_error_msgs.append("Error in ROB (Start) date. Default date of "
                                                   "01-Jan will be used. Please check data;")
                            self.fill_cell(df, index, "Date (DD-Mmm-YYYY)*", error_yellow_fill, "IMO DCS BDN Summary")

                    if rob_start_cnt > 1 and (row["Entry Type*"] == "ROB (Start)"):
                        temp_error_msgs.append("Error due to Multiple ROB (Start) date. Please check data;")
                        self.fill_cell(df, index, "Date (DD-Mmm-YYYY)*", error_yellow_fill, "IMO DCS BDN Summary")

                    if rob_end_cnt == 1 and (row["Entry Type*"] == "ROB (End)"):
                        rob_end_date_val = row["Date (DD-Mmm-YYYY)*"]
                        if isinstance(rob_end_date_val, datetime) and (not self.is_NaT_NaN(rob_end_date_val)):
                            global end_year_date
                            end_year_date = rob_end_date_val + timedelta(hours=23, minutes=59, seconds=59)
                            if EoY_flag:
                                if end_year_date.day != 31 or end_year_date.month != 12:
                                    temp_error_msgs.append("'End of Year' selected but the ROB (End) is not 31-Dec. " \
                                                           "Please check ROB (End) date;")
                                    self.fill_cell(df, index, "Date (DD-Mmm-YYYY)*", error_yellow_fill,
                                                   "IMO DCS BDN Summary")

                            if CoM_flag:
                                if end_year_date.day == 31 and end_year_date.month == 12:
                                    temp_error_msgs.append("'Change in Management / Change in Flag' selected but "
                                                           "the ROB (End) is 31-Dec. Please check ROB (End) date")
                                    self.fill_cell(df, index, "Date (DD-Mmm-YYYY)*", error_yellow_fill,
                                                   "IMO DCS BDN Summary")
                        else:
                            temp_error_msgs.append("Error in ROB (End) date. Default date of "
                                                   "31-Dec will be used. Please check data;")
                            self.fill_cell(df, index, "Date (DD-Mmm-YYYY)*", error_yellow_fill, "IMO DCS BDN Summary")

                    if rob_end_cnt > 1 and (row["Entry Type*"] == "ROB (End)"):
                        temp_error_msgs.append("Error due to Multiple ROB (End) date. Please check data;")
                        self.fill_cell(df, index, "Date (DD-Mmm-YYYY)*", error_yellow_fill, "IMO DCS BDN Summary")

                    row_values = row[2:].values
                    assert_empty_row = [((val =="") or (pd.isnull(val))) for val in row_values]
                    if all(assert_empty_row):
                        temp_error_msgs.append("Empty row for an entry type")
                        self.fill_cell(df, index, "Entry Type*", error_yellow_fill, "IMO DCS BDN Summary")

                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check RoB (Start) and RoB (End) date")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check ROB (Start) and RoB (End) date")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            msg = ("Warning: The default date of {} and end date of {} will be used"
                   "".format(self.date_format(start_year_date), self.date_format(end_year_date)))
            self.message_box(msg)
            dcs_bdn_errors.append(msg)

        '''
        ## Check and get "ROB Correction (Start) and (End)"
        try:
            demo = (datetime.now())
            if dcs_bdn_flag:
                if df["Entry Type*"].str.count(r'(Start)').sum() == 1:
                    index_loc = df[df['Entry Type*'] == "ROB (Start)"].index.values.astype(int)[0]
                    date_index_loc = df.iloc[index_loc]["Date (DD-Mmm-YYYY)*"]
                    if isinstance(date_index_loc, datetime) and (not self.is_NaT_NaN(date_index_loc)):
                        global start_year_date
                        start_year_date = df.iloc[index_loc]["Date (DD-Mmm-YYYY)*"]
                    else:
                        raise
                elif df["Entry Type*"].str.count(r'(Start)').sum() == 0:
                    dcs_bdn_errors.append("Error: The ROB (Start) is not provided")
                    raise
                else:
                    dcs_bdn_errors.append("Error: Multiple ROB (Start) is provided")
                    raise

                if df["Entry Type*"].str.count(r'(End)').sum() == 1:
                    index_loc = df[df['Entry Type*'] == "ROB (End)"].index.values.astype(int)[0]
                    date_index_loc = df.iloc[index_loc]["Date (DD-Mmm-YYYY)*"]
                    if isinstance(date_index_loc, datetime) and (not self.is_NaT_NaN(date_index_loc)):
                        global end_year_date
                        end_year_date = (df.iloc[index_loc]["Date (DD-Mmm-YYYY)*"]) + \
                                        timedelta(days=1) - timedelta(minutes=1)
                        if EoY_flag:
                            if end_year_date.day != 31 or end_year_date.month != 12:
                                temp_error_msgs.append("'End of Year' selected but the ROB (End) is not 31-Dec. " \
                                                       "Please check ROB (End) date")
                                self.fill_cell(df, index_loc, date_colname, error_yellow_fill,
                                               "IMO DCS BDN Summary")
                                temp_msg = (df.iloc[index_loc]["Errors"]).split("\n")
                                temp_msg.extend(temp_error_msgs)
                                temp_msg = list(filter(None, temp_msg))
                                df.at[index_loc, 'Errors'] = "\n".join(temp_msg)

                        if CoM_flag:
                            if end_year_date.day == 31 and end_year_date.month == 12:
                                temp_error_msgs.append("'Change in Management / Chang in Flag' selected but "
                                                       "the ROB (End) is 31-Dec-202x. Please check ROB (End) date")

                                self.fill_cell(df, index_loc, date_colname, error_yellow_fill,
                                               "IMO DCS BDN Summary")
                                temp_msg = (df.iloc[index_loc]["Errors"]).split("\n")
                                temp_msg.extend(temp_error_msgs)
                                temp_msg = list(filter(None, temp_msg))
                                df.at[index_loc, 'Errors'] = "\n".join(temp_msg)

                    else:
                        raise
                elif df["Entry Type*"].str.count(r'(End)').sum() == 0:
                    dcs_bdn_errors.append("Error: The ROB Correction (End) is not provided")
                    raise
                else:
                    dcs_bdn_errors.append("Error: Multiple ROB Correction (End) is provided")
                    raise

                msg = ("OK: Check ROB Correction (Start) and (End)")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check ROB Correction (Start) and (End)")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            msg = ("Warning: The default date of {} and end date of {} will be used"
                   "".format(self.date_format(start_year_date), self.date_format(end_year_date)))
            self.message_box(msg)
            dcs_bdn_errors.append(msg)
        '''

        try:
            ## Check date range
            if dcs_bdn_flag:
                for index, row in df.iterrows():
                    date_value = row[date_colname]
                    temp_error_msgs = []
                    if not (row.isnull().all()):
                        if not self.is_NaT_NaN(date_value):
                            if isinstance(date_value, datetime):
                                if not self.check_date_range(date_value):
                                    temp_error_msgs.append("Please enter date within the current report period"
                                                           " from {} to {};".format(self.date_format(start_year_date),
                                                                                    self.date_format(end_year_date)))
                                    self.fill_cell(df, index, date_colname, error_yellow_fill,
                                                   "IMO DCS BDN Summary")
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check date range")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check date range")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_bdn_errors.append(msg)

        try:
            agg_bdn_fuel = [wb["IMO DCS BDN Summary"].cell(row=6, column=i).value for i in range(3, 12)]
            agg_bdn_fuel = ['%.2f' % elem for elem in agg_bdn_fuel]
            df_analysis.loc["Aggregated BDN Summary"] = agg_bdn_fuel
        except:
            msg = ("Critical Error: Sum of fuel column values of BDN Summary")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            df_analysis_errors.append(msg)
            sum_bdn_data = False

        try:
            if dcs_bdn_flag:
                error_column = ["Errors"]
                error_column.extend(df["Errors"])
                df_errors["IMO DCS BDN Summary"] = df["Errors"]
                self.add_column_openpyxl(error_column, "IMO DCS BDN Summary")
                msg = ("OK: Appending 'Error' column into IMO DCS BDN Summary sheet for file {}".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
                msg = ("OK: Analysis of IMO DCS BDN Summary Successful for file {}".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Appending 'Error' column into IMO DCS BDN Summary sheet for file {}".format(f_name))
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_bdn_errors.append(msg)

        # sys.stdout = old_stdout
        # log_file.close()

        ## =================================================================================================================
        ## For analysing IMO DCS Data
        msg = ("Analysis started for file: {} and sheet name {}".format(f_name, "IMO DCS Data"))
        msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
        self.message_box(msg)
        dcs_data_flag = False
        dcs_data_errors = []
        df = pd.DataFrame()
        try:
            df = pd.read_excel(filename, sheet_name='IMO DCS Data', skiprows=dcs_data_skip_row,
                               usecols=range(dcs_data_end_col), keep_default_na=False,
                               na_values=[''])
            msg = ("OK: Importing sheet of IMO DCS Data")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_flag = True
        except:
            msg = ("Critical Error: Importing sheet of IMO DCS Data")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_flag = False
            dcs_data_errors.append(msg)

        df["Errors"] = ""
        if df.empty:
            dcs_data_errors.append("Empty Data Frame;")

        ## Check the imported file or template
        dcs_missing_columns = []
        for col in dcs_template_columns:
            if col not in df.columns:
                dcs_missing_columns.append(col)

        if len(dcs_missing_columns) != 0:
            msg = ("Critical Error: Template Missing columns: {}".format(",".join(bdn_missing_columns)))
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_flag = False
            dcs_data_errors.append(msg)
        else:
            msg = ("OK: Template")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_flag = True

        try:
            ## Check if numeric in columns
            if dcs_bdn_flag:
                for index, row in df.iterrows():
                    temp_error_msgs = []
                    non_numeric_colnames = []
                    if not (row.isnull().all()):
                        for col_name in dcs_numeric_columns:
                            if not self.is_NaT_NaN(row[col_name]):
                                if not self.is_numeric(row[col_name]):  ## Check if the data is numeric
                                    non_numeric_colnames.append(col_name)
                                    self.fill_cell(df, index, col_name, error_yellow_fill, "IMO DCS Data")
                        if len(non_numeric_colnames) > 0:
                            temp_error_msgs.append("Non-Numeric Value in Column(s): {};"
                                                   .format(",".join(non_numeric_colnames)))
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check if numeric in columns")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check if numeric in columns")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_errors.append(msg)

        try:
            ## Check if non-negative in columns
            if dcs_data_flag:
                for index, row in df.iterrows():
                    temp_error_msgs = []
                    negative_val_colnames = []
                    if not (row.isnull().all()):
                        for col_name in dcs_numeric_columns:
                            if not self.is_NaT_NaN(row[col_name]) and self.is_numeric(row[col_name]):
                                if not self.is_nonnegative(row[col_name]):  ## Check if the data is non-negative
                                    negative_val_colnames.append(col_name)
                                    self.fill_cell(df, index, col_name, error_yellow_fill, "IMO DCS Data")
                        if len(negative_val_colnames) > 0:
                            temp_error_msgs.append("Negative Value in Column(s): {};"
                                                   .format(",".join(negative_val_colnames)))
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check if non-negative in columns")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check if non-negative in columns")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_errors.append(msg)

        try:
            ## Check for non-zero in hours underway, distance travelled and total fuel
            if dcs_data_flag:
                for index, row in df.iterrows():
                    hrs_underway = row["Hours Underway (h)*"]
                    dst_travelled = row["Distance Traveled (nm)*"]
                    total_fuel = np.nansum([val for val in row[dcs_fuel_columns] if isinstance(val, numbers.Number)])
                    temp_error_msgs = []
                    if not (row.isnull().all()):
                        if self.is_numeric(hrs_underway) and self.is_numeric(dst_travelled):
                            if hrs_underway > 0 and (dst_travelled == 0 or self.is_NaT_NaN(dst_travelled)):
                                temp_error_msgs.append(
                                    "'Distance Travelled' is zero while 'Hours Underway' is not. "
                                    "Please check your data;")
                                self.fill_cell(df, index, "Distance Traveled (nm)*", error_yellow_fill, "IMO DCS Data")
                            elif (hrs_underway == 0 or self.is_NaT_NaN(hrs_underway)) and dst_travelled > 0:
                                temp_error_msgs.append(
                                    "'Hours Underway' is zero while 'Distance Travelled' is not. "
                                    "Please check your data;")
                                self.fill_cell(df, index, "Hours Underway (h)*", error_yellow_fill, "IMO DCS Data")
                            if hrs_underway != 0 and dst_travelled != 0:
                                if self.is_numeric(total_fuel):
                                    if total_fuel == 0 or self.is_NaT_NaN(total_fuel):
                                        temp_error_msgs.append(
                                            "Total Fuel is zero while 'Distance Travelled/ Hours Underway' is not. "
                                            "Please check your data;")
                                        self.fill_cell(df, index, "Distance Traveled (nm)*", error_yellow_fill,
                                                       "IMO DCS Data")
                                        self.fill_cell(df, index, "Hours Underway (h)*", error_yellow_fill,
                                                       "IMO DCS Data")
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check for non-zero in hours underway, distance travelled and fuel consumption")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check for non-zero in hours underway, distance travelled and fuel consumption")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_errors.append(msg)

        try:
            ## Check for non-zero in Other Consumption (MT) and Emission Factor
            if dcs_data_flag:
                for index, row in df.iterrows():
                    other_consump = row["Other Consumption (MT)"]
                    other_emission = row["Other Emission Factor (MT CO2/MT fuel)"]
                    temp_error_msgs = []
                    if not (row.isnull().all()):
                        if self.is_numeric(other_consump) and self.is_numeric(other_emission):
                            if abs(other_consump) > 0 and (other_emission == 0 or self.is_NaT_NaN(other_emission)):
                                temp_error_msgs.append(
                                    "When a value for “Other Consumption” is entered, "
                                    "a value for “Other Emission Factor” is to be entered."
                                    "Please check your data;")
                                self.fill_cell(df, index, "Other Consumption (MT)", error_yellow_fill, "IMO DCS Data")
                            elif (other_consump == 0 or self.is_NaT_NaN(other_consump)) and abs(other_emission) > 0:
                                temp_error_msgs.append(
                                    "When a value for “Other Emission Factor” is entered, "
                                    "a value for “Other Consumption” is to be entered."
                                    "Please check your data;")
                                self.fill_cell(df, index, "Other Emission Factor (MT CO2/MT fuel)", error_yellow_fill,
                                               "IMO DCS Data")
                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check for non-zero in Other Consumption (MT) and Emission Factor")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check for non-zero in Other Consumption (MT) and Emission Factor")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_errors.append(msg)
        '''
        try:
            ## Check for Direct CO2 emissions measurement (MT)
            for index, row in df.iterrows():
                co2_emission = row["Direct CO2 emissions measurement (MT), if applicable"]
                non_zero_fuel = []
                temp_error_msgs = []
                total_fuel_usage = 0
                if not (row.isnull().all()):
                    if not self.is_NaT_NaN(co2_emission):
                        for fuel_type in dcs_fuel_columns:
                            if self.is_numeric(row[fuel_type]) and self.is_nonnegative(row[fuel_type]):
                                if row[fuel_type] > 0:
                                    non_zero_fuel.append(fuel_type)
                                    total_fuel_usage = total_fuel_usage + row[fuel_type]

                        if self.is_numeric(co2_emission) and self.is_nonnegative(co2_emission):
                            if co2_emission == 0 and total_fuel_usage != 0:
                                temp_error_msgs.append("CO2 emission is zero while {} consumption is not zero. "
                                                       "Please check your data;".format(",".join(non_zero_fuel)))
                                self.fill_cell(df, index, "Direct CO2 emissions measurement (MT), if applicable",
                                          error_yellow_fill, "IMO DCS Data")

                            elif co2_emission != 0 and total_fuel_usage == 0:
                                temp_error_msgs.append("Total fuel usage is zero while CO2 emission is not zero"
                                                       "Please check your data;")
                                self.fill_cell(df, index, "Direct CO2 emissions measurement (MT), if applicable",
                                          error_yellow_fill, "IMO DCS Data")

                temp_msg = (df.iloc[index]["Errors"]).split("\n")
                temp_msg.extend(temp_error_msgs)
                temp_msg = list(filter(None, temp_msg))
                df.at[index, 'Errors'] = "\n".join(temp_msg)
            msg = ("OK: Check for Direct CO2 emissions measurement (MT)")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " "+ msg)
            self.message_box(msg)
        except:
            msg = ("Error: Check for Direct CO2 emissions measurement (MT)")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " "+ msg)
            self.message_box(msg)
        '''

        start_date_colname = "Start Date and Time (DD-Mmm-YYYY HH:MM) (UTC)*"
        end_date_colname = "End Date and Time (DD-Mmm-YYYY HH:MM) (UTC)*"
        try:
            ## Check for date format and empty date
            if dcs_data_flag:
                for index, row in df.iterrows():
                    temp_error_msgs = []
                    start_date_value = row[start_date_colname]
                    end_date_value = row[end_date_colname]
                    if not (row.isnull().all()):
                        if not isinstance(start_date_value, datetime):
                            temp_error_msgs.append("Error in start date/time format (i.e. DD-Mmm-YYY HH:MM). "
                                                   "Please check data;")
                            self.fill_cell(df, index, start_date_colname, error_yellow_fill, "IMO DCS Data")
                        if isinstance(start_date_value, datetime) and self.is_NaT_NaN(start_date_value):
                            temp_error_msgs.append("Empty cell in start date/time (i.e. DD-Mmm-YYY HH:MM). "
                                                   "Please check data;")
                            self.fill_cell(df, index, start_date_colname, error_yellow_fill, "IMO DCS Data")

                        if not isinstance(end_date_value, datetime):
                            temp_error_msgs.append("Error in end date/time format (i.e. DD-Mmm-YYY HH:MM). "
                                                   "Please check data;")
                            self.fill_cell(df, index, end_date_colname, error_yellow_fill, "IMO DCS Data")
                        if isinstance(end_date_value, datetime) and self.is_NaT_NaN(end_date_value):
                            temp_error_msgs.append("Empty cell in end date/time (i.e. DD-Mmm-YYY HH:MM). "
                                                   "Please check data;")
                            self.fill_cell(df, index, end_date_colname, error_yellow_fill, "IMO DCS Data")

                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check for date format")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check for date format")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_errors.append(msg)

        try:
            ## Check date range
            if dcs_data_flag:
                for index, row in df.iterrows():
                    start_date_value = row[start_date_colname]
                    end_date_value = row[end_date_colname]
                    temp_error_msgs = []
                    if not (row.isnull().all()):
                        if not self.is_NaT_NaN(start_date_value):
                            if isinstance(start_date_value, datetime):
                                if not self.check_date_range(start_date_value):
                                    temp_error_msgs.append("Please enter date within the current report period"
                                                           " from {} to {};".format(self.date_format(start_year_date),
                                                                                    self.date_format(end_year_date)))
                                    self.fill_cell(df, index, start_date_colname, error_yellow_fill, "IMO DCS Data")

                        if not self.is_NaT_NaN(end_date_value):
                            if isinstance(end_date_value, datetime):
                                if not self.check_date_range(end_date_value):
                                    temp_error_msgs.append("Please enter date within the current report period"
                                                           " from {} to {};".format(self.date_format(start_year_date),
                                                                                    self.date_format(end_year_date)))
                                    self.fill_cell(df, index, end_date_colname, error_yellow_fill, "IMO DCS Data")

                        if (not self.is_NaT_NaN(start_date_value)) and (not self.is_NaT_NaN(end_date_value)):
                            if isinstance(start_date_value, datetime) and isinstance(end_date_value, datetime):
                                if not (end_date_value > start_date_value):
                                    temp_error_msgs.append(
                                        "The voyage ending date and time is earlier than the starting "
                                        "date and time. Please enter a valid end date and time;")
                                    self.fill_cell(df, index, end_date_colname, error_yellow_fill, "IMO DCS Data")

                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check date range")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check date range")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_errors.append(msg)
        ############
        ##############
        try:
            ## Check the date flow
            if dcs_data_flag:
                first_row_flag = True
                for index, row in df.iterrows():
                    start_date_value = row[start_date_colname]
                    end_date_value = row[end_date_colname]
                    date_time_assert = isinstance(start_date_value, datetime) and \
                                       isinstance(end_date_value, datetime) and \
                                       not (self.is_NaT_NaN(start_date_value) and \
                                            self.is_NaT_NaN(end_date_value))
                    temp_error_msgs = []
                    if first_row_flag == True:
                        first_row_flag = False
                        if date_time_assert:
                            if self.date_format(start_date_value) != self.date_format(start_year_date):
                                temp_error_msgs.append("Error in start date based on RoB (Start): {};".
                                                       format(self.date_format(start_year_date)))
                                self.fill_cell(df, index, start_date_colname, error_yellow_fill, "IMO DCS Data")
                            prev_end_date_value = end_date_value

                    elif (date_time_assert):
                        prev_date_assert = (not self.is_NaT_NaN(prev_end_date_value)) and \
                                           isinstance(prev_end_date_value, datetime) \
                                           and self.check_date_range(prev_end_date_value)
                        if prev_date_assert:
                            missing_days_cnt = (start_date_value - prev_end_date_value)
                            missing_duration_in_s = missing_days_cnt.total_seconds()
                            missing_days = missing_days_cnt.days
                            missing_hours = (missing_days_cnt.seconds // 3600)
                            missing_minutes = (missing_days_cnt.seconds // 60)
                            missing_minutes = (missing_duration_in_s // 60) % (60*missing_minutes/abs(missing_minutes))
                            tot_missing_minutes = round(divmod(missing_duration_in_s, 60)[0],2)
                            if (tot_missing_minutes > 1):
                                temp_error_msgs.append("The current row's voyage starting time is missing "
                                                       "{} days, {} hours and {} minutes from previous voyage. "
                                                       "Please enter a valid start"
                                                       " time;".format(abs(missing_days), abs(missing_hours),
                                                                      abs(missing_minutes)))
                                self.fill_cell(df, index, start_date_colname, error_yellow_fill, "IMO DCS Data")

                            elif (tot_missing_minutes < 0):
                                temp_error_msgs.append("The current row's voyage starting time is overlapping "
                                                       "{} days, {} hours and {} minutes from previous voyage. "
                                                       "Please enter a valid start"
                                                       " time;".format(abs(missing_days), abs(missing_hours),
                                                                      abs(missing_minutes)))
                                self.fill_cell(df, index, start_date_colname, error_yellow_fill, "IMO DCS Data")
                            prev_end_date_value = end_date_value
                        else:
                            if isinstance(end_date_value, datetime):
                                prev_end_date_value = end_date_value
                    else:
                        if isinstance(end_date_value, datetime):
                            prev_end_date_value = end_date_value

                    if (len(df)-1)==index:
                        if isinstance(end_date_value, datetime):
                            last_row_end_timedelta = end_date_value - end_year_date
                            last_missing_minutes = round(divmod(last_row_end_timedelta.total_seconds(), 60)[0], 2)
                            if abs(last_missing_minutes) > 1:
                                temp_error_msgs.append("Error in end date based on RoB (End): {};".
                                                       format(self.date_format(end_year_date)))
                                self.fill_cell(df, index, end_date_colname, error_yellow_fill, "IMO DCS Data")



                    temp_msg = (df.iloc[index]["Errors"]).split("\n")
                    temp_msg.extend(temp_error_msgs)
                    temp_msg = list(filter(None, temp_msg))
                    df.at[index, 'Errors'] = "\n".join(temp_msg)
                msg = ("OK: Check the date flow")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Check the date flow")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_errors.append(msg)

        try:
            if dcs_data_flag:
                error_column = ["Errors"]
                error_column.extend(df["Errors"])
                df_errors["IMO DCS Data"] = df["Errors"]
                self.add_column_openpyxl(error_column, "IMO DCS Data")
                msg = ("OK: Appending 'Error' column into IMO DCS Data sheet for file {}".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
                msg = ("OK: Analysis of IMO DCS Data Successful for file {}".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
        except:
            msg = ("Error: Appending 'Error' column into IMO DCS BDN Summary sheet for file {}".format(f_name))
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            dcs_data_errors.append(msg)
        # sys.stdout = old_stdout
        # log_file.close()

        try:
            agg_dcs_fuel = [wb["IMO DCS Data"].cell(row=6, column=i).value for i in range(5, 14)]
            agg_dcs_fuel = ['%.2f' % elem for elem in agg_dcs_fuel]
            df_analysis.loc["Aggregated fuel oil consumed"] = agg_dcs_fuel
            sum_dcs_data = True
        except:
            msg = ("Critical Error: Sum of fuel column values of DCS Data")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            df_analysis_errors.append(msg)
            sum_dcs_data = False
        #######################
        ## DF_Analysis
        try:
            df_critical_error = False
            df_analysis = df_analysis.T
            df_analysis["Aggregated BDN Summary"] = df_analysis["Aggregated BDN Summary"].astype(float)
            df_analysis["Aggregated fuel oil consumed"] = df_analysis["Aggregated fuel oil consumed"].astype(float)
            for col in df_analysis.columns:
                sum_value = 0
                for values in df_analysis[col]:
                    try:
                        if not self.is_NaT_NaN(values):
                            sum_value = sum_value + values
                    except:
                        continue
                #print(sum_value)
                df_analysis.loc['Sum Total Fuel', col] = round(sum_value, 2)
            df_analysis['Program check on +/-5% acceptance'] = ((df_analysis['Aggregated fuel oil consumed'] /
                                                                df_analysis['Aggregated BDN Summary'])-1) * 100
            df_analysis["Predicted aggregated fuel oil consumed"] = ""
            df_analysis["% deviation of aggregated fuel oil consumption from prediction"] = ""
            for index, row in df_analysis.iterrows():  ## Limit to two decimals
                try:
                    pct_value = df_analysis.loc[index, 'Program check on +/-5% acceptance']
                    if (pct_value == math.inf) or (pct_value == -math.inf):
                        df_analysis.loc[index, 'Program check on +/-5% acceptance'] = 'infinity'
                    else:
                        df_analysis.at[index, 'Program check on +/-5% acceptance'] = \
                            round(pct_value, 2)
                except:
                    continue
        except:
            msg = ("Error: Data analysis and machine learning prediction")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            df_analysis_errors.append(msg)
            df_critical_error = True

        ## ML Prediction
        ML_errors = []

        try:##
            dst_nm = wb["IMO DCS Data"].cell(row=6, column=4).value
            dst_nm = float(dst_nm)
            if dst_nm is None or (dst_nm == 0):
                raise
            elif not isinstance(dst_nm, (int, float)):
                raise
        except:
            ML_errors.append("Error in obtaining Distance Travelled (Zero or missing or non-numeric)")

        try:##
            hours_uw = wb["IMO DCS Data"].cell(row=6, column=3).value
            hours_uw = float(hours_uw)
            if hours_uw is None or (hours_uw == 0):
                raise
            elif not isinstance(hours_uw, (int, float)):
                raise
        except:
            ML_errors.append("Error in obtaining Hours Underway (Zero or missing or non-numeric)")

        try:##
            me_power = wb['Vessel Details']["B15"].value
            if me_power is None or (me_power == 0):
                raise
            else:
                me_power = [int(num) for num in me_power.split(",")]
                me_power = sum(me_power)
        except:
            ML_errors.append("Error in obtaining Main Engine Power (Zero or missing or non-numeric)")

        try:##
            ae_power = wb['Vessel Details']["B16"].value
            if ae_power is None or (ae_power == 0):
                raise
            else:
                ae_power = [int(num) for num in ae_power.split(",")]
                ae_power = sum(ae_power)
        except:
            ML_errors.append("Error in obtaining Aux Engine Power (Zero or missing or non-numeric)")

        try:##
            gt = wb['Vessel Details']["B12"].value
            if gt is None or (gt == 0):
                raise
            elif not isinstance(gt, (int, float)):
                raise
        except:
            ML_errors.append("Error in obtaining Gross Tonnage (Zero or missing or non-numeric)")
        ################
         ##############
        try:##
            ship_typ = wb['Vessel Details']["B7"].value
            all_types_dict = {"Bulk Carrier": 'Bulk carrier',
                              "Combination Carrier": 'Combination carrier',
                              "Containerships": 'Containership',
                              "Gas Carrier": 'Gas carrier',
                              "General Cargo Ship": 'General cargo ship',
                              "LNG Carrier": 'LNG carrier',
                              "Others": 'Others',
                              "Passenger Ship": 'Passenger ship',
                              "Cruise Passenger Ship": 'Passenger ship',
                              "Refrigerated Cargo Carrier": 'Refrigerated cargo carrier',
                              "Ro-ro Cargo Ship": 'Ro-ro cargo ship',
                              "Ro-ro Passenger Ship": 'Ro-ro cargo ship',
                              "Ro-ro Cargo Ship (Vehicle Carrier)": 'Ro-ro cargo ship (vehicle carrier)',
                              "Tanker": 'Tanker'}

            ML_ship_types = ['Bulk carrier', 'Combination carrier', 'Containership', 'Gas carrier',
                             'General cargo ship', 'LNG carrier', 'Others', 'Passenger ship',
                             'Refrigerated cargo carrier', 'Ro-ro cargo ship',
                             'Ro-ro cargo ship (vehicle carrier)', 'Tanker']

            ship_typ = [1 if all_types_dict[ship_typ] == x else 0 for x in ML_ship_types]
        except:
            ML_errors.append("Error in obtaining Ship Type")

        ########################

        ## Machine Learning Prediction
        try:
            ml_folder = "S:\\Hull\\2021 DCS MRV\\DCS\\ML Trained\\"
            model = load_model(ml_folder + 'DCS model.h5')
            scaler = joblib.load(ml_folder + "DCS_Scalar.save")
            ##'total_distance','total_time','total_main','total_aux','GT','ship_type'
            if len(ML_errors) == 0:
                prediction_data = [dst_nm, hours_uw, me_power, ae_power, gt]
                prediction_data.extend(ship_typ)
                prediction_data = np.array(prediction_data)
                prediction_data = scaler.transform(prediction_data.reshape(1, 17))
                y_hat = model.predict(prediction_data)
                y_hat = copy.deepcopy(round(y_hat[0][0], 2))
                if df_critical_error == False:
                    df_analysis.at["Sum Total Fuel", 'Predicted aggregated fuel oil consumed'] = y_hat
                    max_fuel = max(df_analysis.loc["Sum Total Fuel", "Aggregated fuel oil consumed"],
                                   df_analysis.loc["Sum Total Fuel", "Aggregated BDN Summary"])
                    df_analysis.at["Sum Total Fuel",
                                   "% deviation of aggregated fuel oil consumption from prediction"] = \
                        round(((max_fuel / y_hat) - 1) * 100, 2)
                    #print(round(((max_fuel / y_hat) - 1) * 100, 2))
            else:
                ML_errors.append("Prediction error! Check data")
        except:
            ML_errors.append("Prediction error! Check data")

        try:
            ## Saving 'df_analysis' as 'Program Checklist'
            sheet_name = 'Program Checklist'
            if 'Program Checklist' in wb.sheetnames:  # to check whether sheet you need already exists
                ws = wb['Program Checklist']
            else:
                wb.create_sheet('Program Checklist')
                ws = wb['Program Checklist']

            df_analysis = df_analysis.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)
            df_analysis = df_analysis.astype(str)
            df_analysis = df_analysis.replace('nan', '', regex=True)
            df_analysis = df_analysis.replace('NaT', '', regex=True)
            for r in dataframe_to_rows(df_analysis, index=True, header=True):
                if any(r):
                    ws.append(r)
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            # To make border
            for row_num in range(ws.max_row):
                for col_num in range(ws.max_column):
                    ws.cell(row=row_num + 1, column=col_num + 1).border = thin_border
                    if row_num == 0 or col_num == 0:
                        ws.cell(row=row_num + 1, column=col_num + 1).font = openpyxl.styles.Font(bold=True)
            ## To autofit the columns
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
        except:
            msg = ("Error: Saving sheet 'Program Checklist")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            df_analysis_errors.append(msg)
        ###########################
        ## Analysis of df_analysis [deviation from +- 5% deviation]
        try:
            df_analysis["Errors"] = ""
            for idx, row in enumerate(df_analysis.itertuples(index=False)):
                temp_error_msgs = []
                col_name = "Program check on +/-5% acceptance"
                try:
                    check_pct = float(row[2])
                except:
                    continue
                if not self.is_NaT_NaN(check_pct):
                    if self.is_numeric(check_pct):  ## Check if the data is numeric
                        if abs(check_pct) > 5:
                            error_msg = "WARNING: Wide margin noted between aggregated {} fuel consumption between " \
                                        "'IMO DCS BDN Summary' and 'IMO DCS Data'. " \
                                        "Please double check your data.".format(df_analysis.index[idx])
                            temp_error_msgs.append(error_msg)
                            self.fill_cell(df_analysis, idx, "Program check on +/-5% acceptance",
                                           error_yellow_fill, "Program Checklist")
                temp_msg = (df_analysis.at[df_analysis.index[idx], "Errors"]).split("\n")
                temp_msg.extend(temp_error_msgs)
                temp_msg = list(filter(None, temp_msg))
                df_analysis.at[df_analysis.index[idx], 'Errors'] = "\n".join(temp_msg)
            error_column = ["Errors"]
            error_column.extend(df_analysis["Errors"])
            self.add_column_openpyxl(error_column, "Program Checklist")
            msg = ("OK: Analysis of 5% deviation")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
        except:
            msg = ("Error: Analysis of 5% deviation")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            df_analysis_errors.append(msg)

        try:
            if (len(vessel_details_errors) != 0) or (not all(df_errors['Vessel Details'] == '')):
                vessel_details_errors = "Please check the errors as highlighted"
                wb["Vessel Details"].sheet_properties.tabColor = "00FFFF00"
            else:
                vessel_details_errors = "Analysis completed without errors"
                wb["Vessel Details"].sheet_properties.tabColor = "0000FF00"

            if (len(report_setup_errors) != 0) or (not all(df_errors['Report Setup'] == '')):
                report_setup_errors = "Please check the errors in error column"
                wb["Report Setup"].sheet_properties.tabColor = "00FFFF00"
            else:
                report_setup_errors = "Analysis completed without errors"
                wb["Report Setup"].sheet_properties.tabColor = "0000FF00"

            if len(dcs_bdn_errors) != 0 or (not all(df_errors['IMO DCS BDN Summary'] == '')):
                dcs_bdn_errors = "Please check the errors in error column"
                wb["IMO DCS BDN Summary"].sheet_properties.tabColor = "00FFFF00"
            else:
                dcs_bdn_errors = "Analysis completed without errors"
                wb["IMO DCS BDN Summary"].sheet_properties.tabColor = "0000FF00"

            if len(dcs_data_errors) != 0 or (not all(df_errors['IMO DCS Data'] == '')):
                dcs_data_errors = "Please check the errors in error column"
                wb["IMO DCS Data"].sheet_properties.tabColor = "00FFFF00"
            else:
                dcs_data_errors = "Analysis completed without errors"
                wb["IMO DCS Data"].sheet_properties.tabColor = "0000FF00"

            if len(df_analysis_errors) != 0 or len(ML_errors) != 0:
                df_analysis_errors.extend(ML_errors)
                df_analysis_errors = ";\n".join(df_analysis_errors)
                wb["Program Checklist"].sheet_properties.tabColor = "00FFFF00"
            else:
                df_analysis_errors = "Analysis completed without errors"
                wb["Program Checklist"].sheet_properties.tabColor = "0000FF00"

            wb.create_sheet('Error Summary')
            ws = wb['Error Summary']
            ws['A1'] = "Vessel Details"
            ws['B1'] = vessel_details_errors
            ws['A2'] = "Report Setup"
            ws['B2'] = report_setup_errors
            ws['A3'] = "IMO DCS BDN Summary"
            ws['B3'] = dcs_bdn_errors
            ws['A4'] = "IMO DCS Data"
            ws['B4'] = dcs_data_errors
            ws['A5'] = "Program Checklist"
            ws['B5'] = df_analysis_errors
            msg = ("OK: Created sheet Error Summary")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
        except:
            msg = ("Error: Creating sheet Error Summary")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)

    def sampling_data(self, source_dir, f_name):
        ## ==============================================================================================================
        ## Sampling data
        msg = ("Sampling data for file: {} and sheet name {}".format(f_name, "IMO DCS Data"))
        msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
        self.message_box(msg)
        dcs_data_errors = []
        df = pd.DataFrame()
        filename = source_dir + "\\" + f_name
        try:
            df = pd.read_excel(filename, sheet_name='IMO DCS Data', skiprows=dcs_data_skip_row,
                               usecols=range(dcs_data_end_col), keep_default_na=False, na_values=[''])
            msg = ("OK: Importing sheet of IMO DCS Data")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)

        except:
            msg = ("Critical Error: Importing sheet of IMO DCS Data")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)

            dcs_data_errors.append(msg)

        if len(df) <= 12:
            pct_sample = round(0.75 * len(df))
        elif len(df) <= 30:
            pct_sample = round(0.50 * len(df))
        elif len(df) <= 50:
            pct_sample = round(0.34 * len(df))
        elif len(df) <= 90:
            pct_sample = round(0.21 * len(df))
        elif len(df) <= 150:
            pct_sample = round(0.14 * len(df))
        elif len(df) <= 280:
            pct_sample = round(0.09 * len(df))
        elif len(df) <= 500:
            pct_sample = round(0.05 * len(df))

        sample_factor = 1
        try:
            sheet_name = 'Program Checklist'
            ws = wb[sheet_name]
            #print(abs(float(ws["F11"].value)))
            if abs(float(ws["F11"].value)) <= 20:
                sample_factor = 2
        except:
            msg = ("No ML value available! Default sampling selected!")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)

        pct_sample = int(pct_sample / sample_factor)
        chosen_idx = np.random.choice(len(df), replace=False, size=pct_sample)
        sample_df = df.iloc[chosen_idx]
        sample_df.index += 8
        #sample_df.sort_index(inplace=True)
        start_date_name = 'Start Date and Time (DD-Mmm-YYYY HH:MM) (UTC)*'
        end_date_name = 'End Date and Time (DD-Mmm-YYYY HH:MM) (UTC)*'

        for index, row in sample_df.iterrows():
            try:
                sample_df[start_date_name] = sample_df[start_date_name].dt.strftime("%d-%b-%Y %H:%M")
                #sample_df.at[index, start_date_name] = str((datetime.strftime(row[start_date_name], "%d-%b-%Y %H:%M")))
            except:
                continue
            try:
                sample_df[end_date_name] = sample_df[end_date_name].dt.strftime("%d-%b-%Y %H:%M")
                #sample_df.at[index, end_date_name] = (datetime.strftime(row[end_date_name], "%d-%b-%Y  %I:%M:%S %p"))
            except:
                continue
        sample_df.sort_index(inplace=True)

        #################

        ###############
        try:
            ## Saving 'df_analysis' as 'Program Checklist'
            sheet_name = 'Sampling Data'
            if 'Sampling Data' in wb.sheetnames:  # to check whether sheet you need already exists
                ws = wb['Sampling Data']
            else:
                wb.create_sheet('Sampling Data')
                ws = wb['Sampling Data']
            sample_df = sample_df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)
            sample_df = sample_df.astype(str)
            sample_df = sample_df.replace('nan', '', regex=True)
            sample_df = sample_df.replace('NaT', '', regex=True)
            for r in dataframe_to_rows(sample_df, index=True, header=True):
                ws.append(r)

            ws.delete_rows(2, 1)
            msg = ("Sampling data Created!")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
        except:
            msg = ("Error: Saving sheet 'Program Checklist")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)

    def make_GISIS_xlsx(self, source_dir, destination_dir, file_names):
        ## ==============================================================================================================
        msg = ("GISIS data extraction started")
        msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
        self.message_box(msg)
        files = file_names
        col_names_gisis = ["ReportingStartDate", "ReportingEndDate", "ShipFlag", "ShipIMONumber", "ShipType",
                     "ShipTypeOther", "ShipGrossTonnage", "ShipNetTonnage", "ShipDeadweight",
                     "ShipMainPropulsionPowers", "ShipAuxiliaryEnginesPowers", "ShipEEDI", "ShipIceClass",
                     "DistanceTravelled", "HoursUnderway", "ConsumptionData1", "ConsumptionData2", "ConsumptionData3",
                     "ConsumptionData4", "ConsumptionData5", "ConsumptionData6", "ConsumptionData7", "ConsumptionData8",
                     "ConsumptionData9"]
        col_names_soc = ['date1', 'date2', 'date3', 'date4', 'date5', 'date of issue', 'Name of Ship',
                        'Distinctive number or letters', 'IMO number', 'Port of registry', 'Gross tonnage',
                        'place of issue of Statement', 'signature of duly authorized official issuing the statement',
                        'Title', 'full designation of the Party', 'Certificate Number']
        df = pd.DataFrame(columns=col_names_gisis)
        df_soc = pd.DataFrame(columns=col_names_soc)
        unsuccess = pd.DataFrame(columns=["Unsuccess"])
        errors_GISIS = pd.DataFrame(columns=["Reference","Errors"])

        for f_name in files:
            try:
                filename = source_dir + "\\" + f_name
                w_book = openpyxl.load_workbook(filename, data_only=True)
                all_values = []
                error_messages = []
                soc_dict = {i: "" for i in col_names_soc}
                try:
                    all_values.append(w_book['IMO DCS BDN Summary']["B8"].value)  # "ReportingStartDate"
                except:
                    error_messages.append("Error in ReportingStartDate")
                    all_values.append("")
                try:
                    all_values.append(w_book['IMO DCS BDN Summary']["B9"].value)  # "ReportingEndDate"
                except:
                    error_messages.append("Error in ReportingEndDate")
                    all_values.append("")
                try:
                    ship_flag = w_book['Vessel Details']["B9"].value # ShipFlag
                    ship_flag_code = [val for key, val in country_codes.items() if ship_flag.lower() in key.lower()]
                    if len(ship_flag_code) == 1:
                        all_values.append(ship_flag_code[0])
                        soc_dict["full designation of the Party"] = \
                            list(country_codes.keys())[list(country_codes.values()).index(ship_flag_code[0])]
                    else:
                        raise
                except:
                    error_messages.append("Error in ship flag code.")
                    all_values.append("")
                try:
                    IMONumber = w_book['Vessel Details']["B5"].value # IMO Number
                    if len(str(IMONumber)) == 7:
                        all_values.append(IMONumber)
                        soc_dict["IMO number"] = IMONumber
                    else:
                        raise
                except:
                    error_messages.append("Error in IMO Number.")
                    all_values.append("")
                try:
                    ShipType = (w_book['Vessel Details']["B7"].value).replace(" ", "") # ShipType
                    IMO_ShipType = ["BulkCarrier", "GasCarrier", "Tanker", "Containership", "GeneralCargoShip",
                                    "RefrigeratedCargoCarrier", "CombinationCarrier", "PassengerShip",
                                    "RoRoCargoShipVehicleCarrier", "RoRoCargoShip", "RoRoPassengerShip", "LNGCarrier",
                                    "CruisePassengerShip", "Other"]
                    shiptype_matched = [s for s in IMO_ShipType if (ShipType.lower() in s.lower() or s.lower() in ShipType.lower())]
                    if len(shiptype_matched) == 1:
                        all_values.append(shiptype_matched[0])
                    else:
                        raise
                except:
                    error_messages.append("Error in Ship Type")
                    all_values.append("")
                try:
                    ShipTypeOther = w_book['Vessel Details']["B8"].value # ShipTypeOther
                    if shiptype_matched[0] == "Other":
                        if ShipTypeOther == "" or ShipTypeOther is None:
                            raise
                        else:
                            all_values.append(ShipTypeOther)
                    else:
                        all_values.append(ShipTypeOther)
                except:
                    error_messages.append("Error in Ship type other")
                    all_values.append("")
                try:
                    ShipGrossTonnage = w_book['Vessel Details']["B12"].value  # ShipGrossTonnage
                    if float(ShipGrossTonnage) >=5000:
                        all_values.append(int(ShipGrossTonnage))
                        soc_dict["Gross tonnage"] = int(ShipGrossTonnage)
                    else:
                        raise
                except:
                    error_messages.append("Error in Ship Gross Tonnage")
                    all_values.append("")
                try:
                    ShipNetTonnage = w_book['Vessel Details']["B13"].value # ShipNetTonnage
                    if float(ShipNetTonnage) > 0:
                        all_values.append(int(ShipNetTonnage))
                    else:
                        raise
                except:
                    error_messages.append("Error in ship net tonnage")
                    all_values.append("")
                try:
                    ShipDeadweight = w_book['Vessel Details']["B14"].value # ShipDeadweight
                    if float(ShipDeadweight) > 0:
                        all_values.append(int(ShipDeadweight))
                    else:
                        raise
                except:
                    error_messages.append("Error in ship dead weight")
                    all_values.append("")
                try:
                    all_values.append(w_book['Vessel Details']["B15"].value)  # ShipMainPropulsionPowers
                except:
                    error_messages.append("Error in ship main propulsion power")
                    all_values.append("")
                try:
                    all_values.append(w_book['Vessel Details']["B16"].value)  # ShipAuxiliaryEnginesPowers
                except:
                    error_messages.append("Error in ship auxiliary engine power")
                    all_values.append("")
                try:
                    ShipEEDI = w_book['Vessel Details']["B17"].value
                    if float(ShipEEDI) >= 0:
                        all_values.append(ShipEEDI)  # ShipEEDI
                except:
                    error_messages.append("Error in ShipEEDI")
                    all_values.append("")
                try:
                    all_values.append(w_book['Vessel Details']["B18"].value)  # ShipIceClass
                except:
                    error_messages.append("Error in Ship Ice Class")
                    all_values.append("")
                try:
                    DistanceTravelled = w_book['IMO DCS Data']["D6"].value # DistanceTravelled
                    if float(DistanceTravelled) > 0:
                        all_values.append(int(DistanceTravelled))
                    else:
                        raise
                except:
                    error_messages.append("Error in Distance travelled (Cell D6)")
                    all_values.append("")
                try:
                    HoursUnderway = w_book['IMO DCS Data']["C6"].value # HoursUnderway
                    if float(HoursUnderway) > 0:
                        all_values.append(int(HoursUnderway))
                    else:
                        raise
                except:
                    error_messages.append("Error in Hours underway (Cell C6)")
                    all_values.append("")
                try:
                    cm = (w_book['Vessel Details']["B19"].value).replace(" ", "")
                    IMO_methods = ["BDN", "FlowMeter", "BunkerTankMonitoring"]
                    collect_method = [s for s in IMO_methods if s.lower() in cm.lower()]
                    collect_method = collect_method[0]
                except:
                    error_messages.append("Error in collection method")
                    all_values.append("")
                temp_list = []
                if float(w_book['IMO DCS Data']["E6"].value) > 0:
                    temp_list.append(round(float(w_book['IMO DCS Data']["E6"].value), 2))
                    temp_list.append("HeavyFuel")
                    temp_list.append("")
                    temp_list.append(3.114)
                    temp_list.append(collect_method)
                    all_values.append((','.join(str(v) for v in temp_list)))
                temp_list = []
                if float(w_book['IMO DCS Data']["F6"].value) > 0:
                    temp_list.append(round(float(w_book['IMO DCS Data']["F6"].value), 2))
                    temp_list.append("LightFuel")
                    temp_list.append("")
                    temp_list.append(3.151)
                    temp_list.append(collect_method)
                    all_values.append((','.join(str(v) for v in temp_list)))
                temp_list = []
                if float(w_book['IMO DCS Data']["G6"].value) > 0:
                    temp_list.append(round(float(w_book['IMO DCS Data']["G6"].value), 2))
                    temp_list.append("DieselGasOil")
                    temp_list.append("")
                    temp_list.append(3.206)
                    temp_list.append(collect_method)
                    all_values.append((','.join(str(v) for v in temp_list)))
                temp_list = []
                if float(w_book['IMO DCS Data']["H6"].value) > 0:
                    temp_list.append(round(float(w_book['IMO DCS Data']["H6"].value),2))
                    temp_list.append("LPGPropane")
                    temp_list.append("")
                    temp_list.append(3.000)
                    temp_list.append(collect_method)
                    all_values.append((','.join(str(v) for v in temp_list)))
                temp_list = []
                if float(w_book['IMO DCS Data']["I6"].value) > 0:
                    temp_list.append(round(float(w_book['IMO DCS Data']["I6"].value), 2))
                    temp_list.append("LPGButane")
                    temp_list.append("")
                    temp_list.append(3.030)
                    temp_list.append(collect_method)
                    all_values.append((','.join(str(v) for v in temp_list)))
                temp_list = []
                if float(w_book['IMO DCS Data']["J6"].value) > 0:
                    temp_list.append(round(float(w_book['IMO DCS Data']["J6"].value), 2))
                    temp_list.append("LNG")
                    temp_list.append("")
                    temp_list.append(2.750)
                    temp_list.append(collect_method)
                    all_values.append((','.join(str(v) for v in temp_list)))
                temp_list = []
                if float(w_book['IMO DCS Data']["K6"].value) > 0:
                    temp_list.append(round(float(w_book['IMO DCS Data']["K6"].value),2))
                    temp_list.append("Methanol")
                    temp_list.append("")
                    temp_list.append(1.375)
                    temp_list.append(collect_method)
                    all_values.append((','.join(str(v) for v in temp_list)))
                temp_list = []
                if float(w_book['IMO DCS Data']["L6"].value) > 0:
                    temp_list.append(round(float(w_book['IMO DCS Data']["L6"].value),2))
                    temp_list.append("Ethanol")
                    temp_list.append("")
                    temp_list.append(1.913)
                    temp_list.append(collect_method)
                    all_values.append((','.join(str(v) for v in temp_list)))
                temp_list = []
                if float(w_book['IMO DCS Data']["M6"].value) > 0:
                    temp_list.append(round(float(w_book['IMO DCS Data']["M6"].value),2))
                    temp_list.append("Other")
                    temp_list.append("OtherFuelType")
                    temp_list.append(1)
                    temp_list.append(collect_method)
                    all_values.append((','.join(str(v) for v in temp_list)))

                while (len(all_values) != len(col_names_gisis)):
                    all_values.append("")
                soc_dict["Port of registry"] = w_book['Vessel Details']["B10"].value
                soc_dict["Name of Ship"] = w_book['Vessel Details']["B6"].value
                df.loc[len(df)] = all_values
                df_soc = df_soc.append(soc_dict, ignore_index=True)
                print(df_soc)
                errors_GISIS.loc[len(df), "Reference"] = len(df)
                errors_GISIS.loc[len(df), "Errors"] = ";\n".join(error_messages)
            except:
                unsuccess.loc[len(unsuccess)] = f_name

        df.index += 1
        df.index.name = "ReferenceCode"
        df_soc.index +=1
        df_soc.index.name = "ReferenceCode"
        # Creating Excel Writer Object from Pandas
        writer = pd.ExcelWriter(destination_dir + "\\" + 'Output_GISIS_data_xml.xlsx', engine='xlsxwriter')
        workbook = writer.book
        df.to_excel(writer, sheet_name='GISIS Data', startrow=0, startcol=0)
        unsuccess.to_excel(writer, sheet_name='Unsuccessful files', startrow=0, startcol=0)
        errors_GISIS.to_excel(writer, sheet_name='Errors', startrow=0, startcol=0)
        df_soc.to_excel(writer, sheet_name = "SoC Data", startrow=0, startcol=0)
        writer.save()

        msg = ("GISIS data generated for generating xml")
        msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
        self.message_box(msg)

    def misstatement_sampling(self, source_dir, destination_dir, file_names):
        files = file_names
        for f_name in files:
            ## Logger
            #orig_stdout = sys.stdout
            #log_file = open(destination_dir + "\\" + "log_" + f_name + ".log", 'w')
            #sys.stdout = log_file
            filename = source_dir + "\\" + f_name
            ## Import file
            try:
                global wb
                wb = openpyxl.load_workbook(filename, data_only=True)
                msg = ("OK: Loading of file {}.".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
                msg = "==============================================================================="
                self.message_box(msg)
            except:
                msg = ("Error: Loading of file {}.".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
                break

            if self.mistatement.get():
                self.misstatement_prediction(source_dir, f_name)
            if self.sampling.get():
                self.sampling_data(source_dir, f_name)

            #sys.stdout = orig_stdout
            #log_file.close()

            try:
                for sheet_name in wb.sheetnames:  # to check whether sheet you need already exists
                    ws = wb[sheet_name]
                    ws.protection.sheet = False
                wb.save(destination_dir + "\\" + "Output_" + f_name)
                msg = ("OK: Save file {}".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
                msg = "==============================================================================="
                self.message_box(msg)
            except:
                msg = ("Error: Save file {}. Please check if the file is open".format(f_name))
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
    ##### Code for SOC Generation
    def _getFields(self, obj, tree=None, retval=None, fileobj=None):
        fieldAttributes = {'/FT': 'Field Type', '/Parent': 'Parent', '/T': 'Field Name', '/TU': 'Alternate Field Name',
                           '/TM': 'Mapping Name', '/Ff': 'Field Flags', '/V': 'Value', '/DV': 'Default Value'}
        if retval is None:
            retval = OrderedDict()
            catalog = obj.trailer["/Root"]
            # get the AcroForm tree
            if "/AcroForm" in catalog:
                tree = catalog["/AcroForm"]
            else:
                return None
        if tree is None:
            return retval

        obj._checkKids(tree, retval, fileobj)
        for attr in fieldAttributes:
            if attr in tree:
                # Tree is a field
                obj._buildField(tree, retval, fileobj, fieldAttributes)
                break

        if "/Fields" in tree:
            fields = tree["/Fields"]
            for f in fields:
                field = f.getObject()
                obj._buildField(field, retval, fileobj, fieldAttributes)

        return retval

    def get_form_fields(self, infile):
        infile = PdfFileReader(open(infile, 'rb'))
        fields = self._getFields(infile)
        return OrderedDict((k, v.get('/V', '')) for k, v in fields.items())

    def set_need_appearances_writer(self, writer):
        # See 12.7.2 and 7.7.2 for more information:
        # http://www.adobe.com/content/dam/acom/en/devnet/acrobat/pdfs/PDF32000_2008.pdf
        try:
            catalog = writer._root_object
            # get the AcroForm tree and add "/NeedAppearances attribute
            if "/AcroForm" not in catalog:
                writer._root_object.update({
                    NameObject("/AcroForm"): IndirectObject(len(writer._objects), 0, writer)})

            need_appearances = NameObject("/NeedAppearances")
            writer._root_object["/AcroForm"][need_appearances] = BooleanObject(True)
            return writer

        except Exception as e:
            print('set_need_appearances_writer() catch : ', repr(e))
            return writer

    def make_SOC_RR(self,source_dir, destination_dir):

        if self.SOC.get():
            msg = ("SOC Generation started!")
            msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
            self.message_box(msg)
            try:
                filename = source_dir + "\\" + "Output_GISIS_data_xml.xlsx"
                df_soc = pd.read_excel(filename, sheet_name="SoC Data")
                soc_flag = 1
            except:
                msg = ("Error: Missing file Output_GISIS_data_xml.xlsx or sheet 'SoC Data'")
                msg = (str(datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")) + " " + msg)
                self.message_box(msg)
                soc_flag = 0

            if soc_flag:
                df_doc_dict = df_soc.to_dict(orient='records')
                for dict_item in df_doc_dict:
                    infile = 'SOC_Template.pdf'

                    pprint(self.get_form_fields(infile))

                    inputStream = open(infile, "rb")
                    pdf_reader = PdfFileReader(inputStream, strict=False)
                    if "/AcroForm" in pdf_reader.trailer["/Root"]:
                        pdf_reader.trailer["/Root"]["/AcroForm"].update(
                            {NameObject("/NeedAppearances"): BooleanObject(True)})

                    pdf_writer = PdfFileWriter()
                    self.set_need_appearances_writer(pdf_writer)
                    if "/AcroForm" in pdf_writer._root_object:
                        pdf_writer._root_object["/AcroForm"].update(
                            {NameObject("/NeedAppearances"): BooleanObject(True)})

                    field_dictionary = dict_item
                    for key, val in field_dictionary.items():
                        try:
                            if math.isnan(val):
                                field_dictionary[key] = ""
                        except:
                            continue

                    pdf_writer.addPage(pdf_reader.getPage(0))
                    pdf_writer.updatePageFormFieldValues(pdf_writer.getPage(0), field_dictionary)
                    out_filename = field_dictionary["Name of Ship"]
                    outputStream = open(destination_dir +"\\"+out_filename+".pdf", "wb")
                    pdf_writer.write(outputStream)

                    inputStream.close()
                    outputStream.close()
            else:
                return


    def main_program(self, s_dir, d_dir, file_names):
        self.btn_Start.config(state=tk.DISABLED)
        source_dir = s_dir
        destination_dir = d_dir
        files = file_names
        if (self.mistatement.get()) or (self.sampling.get()):
            self.misstatement_sampling(source_dir=source_dir,destination_dir=destination_dir, file_names=files)

        if self.GISIS.get():
            self.make_GISIS_xlsx(source_dir=source_dir,destination_dir=destination_dir, file_names=files)

        if self.SOC.get() or self.RR.get():
            self.make_SOC_RR(source_dir=source_dir,destination_dir=destination_dir)


        self.btn_Start.config(state=tk.NORMAL)
        msg = ("""
=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=
                Completed Analysis! Check your output folder
=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=x=
""")
        self.message_box(msg)


app = IMO_DCS_App()
app.mainloop()